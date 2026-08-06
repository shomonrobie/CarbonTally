"""Workbook discovery and raw-row reading for DEFRA flat-format files.

Discovery is entirely content-driven: no worksheet name is hardcoded. A data
worksheet is recognised by its header row (``ID`` / ``Scope`` / ``GHG/Unit`` /
``GHG Conversion Factor <year>``); metadata sheets are recognised by their
label/value pairs. The reporting year is parsed from the factor column header.
"""
from __future__ import annotations

import hashlib
import os
import re
from typing import Any, Iterable, Iterator, Optional

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .models import RawRow, WorkbookAnalysis, WorkbookMeta, WorksheetInfo

# Header signatures used to classify worksheets (column *labels*, never names).
DATA_HEADER_SIGNATURE: tuple[str, ...] = ("ID", "Scope", "GHG/Unit")
FACTOR_HEADER_RE = re.compile(r"^GHG Conversion Factor\s+(\d{4})$")
META_LABEL_SUFFIXES: tuple[str, ...] = (
    "Year:",
    "Version:",
    "Status:",
    "Next publication date:",
    "Factor set:",
    "Produced by",
    "For technical queries",
)
EMPTY = ""
NBSP = "\u00a0"
MAX_HEADER_SCAN_ROWS = 40
MAX_META_SCAN_ROWS = 80
MAX_META_SCAN_COLS = 16


# ---------------------------------------------------------------------------
# Small text helpers
# ---------------------------------------------------------------------------
def clean_cell(value: Any) -> str:
    """Return a stripped, whitespace-normalised string for a cell value."""
    if value is None:
        return EMPTY
    text = str(value).replace(NBSP, " ").replace("\xa0", " ")
    return re.sub(r"\s+", " ", text).strip()


# ---------------------------------------------------------------------------
# Workbook-level helpers
# ---------------------------------------------------------------------------
def open_workbook(path: str, data_only: bool = True) -> Workbook:
    """Open an xlsx workbook in read-only, values-only mode."""
    if not os.path.isfile(path):
        raise FileNotFoundError(f"Workbook not found: {path}")
    return load_workbook(path, read_only=True, data_only=data_only)


def workbook_sha256(path: str) -> str:
    """SHA-256 of the workbook file for provenance."""
    digest = hashlib.sha256()
    with open(path, "rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 256), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _row_is_data_header(cells: Iterable[Any]) -> bool:
    """True when a row carries the emission-factor header signature."""
    values = [clean_cell(c) for c in cells]
    return all(v in values for v in DATA_HEADER_SIGNATURE) and any(
        FACTOR_HEADER_RE.match(v) for v in values
    )


def detect_header_row(ws: Worksheet) -> Optional[tuple[int, dict[str, int]]]:
    """Locate the data header row and build a {column label: index} map.

    Returns ``(row_index_1based, column_map)`` or ``None``.
    """
    max_col = ws.max_column or 0
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=min(MAX_HEADER_SCAN_ROWS, ws.max_row or 1), values_only=True),
        start=1,
    ):
        cells = tuple(row)
        if not _row_is_data_header(cells[:max_col]):
            continue
        column_map: dict[str, int] = {}
        for col_idx, value in enumerate(cells[:max_col], start=1):
            label = clean_cell(value)
            if label:
                column_map[label] = col_idx
        return row_idx, column_map
    return None


def _sheet_looks_like_metadata(ws: Worksheet) -> bool:
    """A worksheet is metadata when it carries label/value pairs but no data header."""
    if detect_header_row(ws):
        return False
    for row in ws.iter_rows(max_row=min(MAX_META_SCAN_ROWS, ws.max_row or 1), max_col=MAX_META_SCAN_COLS, values_only=True):
        values = [clean_cell(c) for c in row]
        for value in values:
            if any(value.startswith(suffix) for suffix in META_LABEL_SUFFIXES):
                return True
    return False


def scan_metadata_values(ws: Worksheet) -> dict[str, str]:
    """Collect label/value pairs from a metadata worksheet."""
    found: dict[str, str] = {}
    for row in ws.iter_rows(
        max_row=min(MAX_META_SCAN_ROWS, ws.max_row or 1),
        max_col=MAX_META_SCAN_COLS,
        values_only=True,
    ):
        values = [clean_cell(c) for c in row]
        for idx, value in enumerate(values):
            for suffix in META_LABEL_SUFFIXES:
                if not value.startswith(suffix):
                    continue
                label = suffix.rstrip(":").lower().replace(" ", "_")
                if suffix == "Produced by":
                    label = "producer"
                if suffix == "For technical queries":
                    label = "contact"
                # The value follows the label in the same or a later cell.
                following = [v for v in values[idx + 1 :] if v]
                if following:
                    found[label] = following[0]
                break
    return found


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------
def analyze_workbook(path: str) -> tuple[Workbook, WorkbookAnalysis]:
    """Open a workbook and produce its structural analysis.

    Returns ``(workbook, analysis)``. The analysis lists every worksheet, its
    type, detected header/columns and the resolved reporting year.
    """
    wb = open_workbook(path)
    sheet_infos: list[WorksheetInfo] = []
    data_sheet_names: list[str] = []
    reporting_year: Optional[int] = None
    meta_values: dict[str, str] = {}

    for ws in wb.worksheets:
        header = detect_header_row(ws)
        max_row = ws.max_row
        max_col = ws.max_column
        if header is not None:
            row_idx, column_map = header
            year = parse_reporting_year(column_map)
            info = WorksheetInfo(
                name=ws.title,
                sheet_type="data",
                max_row=max_row,
                max_col=max_col,
                header_row=row_idx,
                columns=tuple(sorted(column_map.items(), key=lambda item: item[1])),
                data_row_count=max(0, (max_row or 0) - row_idx),
            )
            sheet_infos.append(info)
            data_sheet_names.append(ws.title)
            if year is not None and reporting_year is None:
                reporting_year = year
        elif _sheet_looks_like_metadata(ws):
            sheet_infos.append(
                WorksheetInfo(
                    name=ws.title,
                    sheet_type="metadata",
                    max_row=max_row,
                    max_col=max_col,
                )
            )
            meta_values.update(scan_metadata_values(ws))
        else:
            sheet_infos.append(
                WorksheetInfo(
                    name=ws.title,
                    sheet_type="other",
                    max_row=max_row,
                    max_col=max_col,
                )
            )

    year_meta: Optional[int] = None
    raw_year = meta_values.get("year", "")
    if raw_year.isdigit():
        year_meta = int(raw_year)

    workbook_meta = WorkbookMeta(
        source_path=path,
        file_sha256=workbook_sha256(path) if os.path.isfile(path) else "",
        file_size_bytes=os.path.getsize(path) if os.path.isfile(path) else None,
        title=meta_values.get("title", ""),
        version=meta_values.get("version", ""),
        year=year_meta,
        status=meta_values.get("status", ""),
        next_publication_date=meta_values.get("next_publication_date", ""),
        factor_set_label=meta_values.get("factor_set", ""),
        producer=meta_values.get("producer", ""),
        contact=meta_values.get("contact", ""),
    )

    if reporting_year is None and year_meta is not None:
        reporting_year = year_meta

    analysis = WorkbookAnalysis(
        meta=workbook_meta,
        worksheets=tuple(sheet_infos),
        data_sheet_names=tuple(data_sheet_names),
        reporting_year=reporting_year,
    )
    return wb, analysis


def iter_data_rows(ws: Worksheet, info: WorksheetInfo) -> Iterator[RawRow]:
    """Yield every worksheet row below the header as a ``RawRow``."""
    if info.header_row is None:
        return
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=info.header_row + 1, max_col=info.max_col or 0, values_only=True),
        start=info.header_row + 1,
    ):
        yield RawRow(sheet_name=info.name, row_number=row_idx, cells=tuple(row))


def parse_reporting_year(column_map: dict[str, int]) -> Optional[int]:
    """Extract the reporting year from the factor-column header."""
    for label in column_map:
        match = FACTOR_HEADER_RE.match(label)
        if match:
            return int(match.group(1))
    return None
