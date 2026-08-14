"""Workbook parser: discovery, classification and row extraction.

Uses openpyxl for exact cell-level reading and pandas for sheet statistics.
No worksheet name is hardcoded: a data worksheet is recognised by its header
signature (ID / Scope / GHG/Unit plus a ``GHG Conversion Factor <year>``
column); documentation sheets by their label/value pairs; anything else is
reported as unsupported (never silently skipped).
"""
from __future__ import annotations

import hashlib
import os
import re
from typing import Any, Iterator, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .models import (
    SHEET_DATA,
    SHEET_DOCUMENTATION,
    SHEET_UNSUPPORTED,
    ParsedRow,
    WorkbookAnalysis,
    WorkbookMeta,
    WorksheetInfo,
)

DATA_HEADER_SIGNATURE: tuple[str, ...] = ("ID", "Scope", "GHG/Unit")
FACTOR_HEADER_RE = re.compile(r"^GHG Conversion Factor\s+(\d{4})$")
META_LABEL_SUFFIXES: tuple[str, ...] = (
    "Year:",
    "Version:",
    "Status:",
    "Next publication date:",
    "Factor set:",
    "Produced by",
    "For technical queries",
)
NBSP = "\u00a0"
MAX_HEADER_SCAN_ROWS = 40
MAX_META_SCAN_ROWS = 80
MAX_META_SCAN_COLS = 16


# ---------------------------------------------------------------------------
# Cell helpers
# ---------------------------------------------------------------------------
def clean_cell(value: Any) -> str:
    """Return a stripped, whitespace-normalised string for a cell value."""
    if value is None:
        return ""
    text = str(value).replace(NBSP, " ").replace("\xa0", " ")
    return re.sub(r"\s+", " ", text).strip()


# ---------------------------------------------------------------------------
# Workbook-level helpers
# ---------------------------------------------------------------------------
def open_workbook(path: str, data_only: bool = True) -> Workbook:
    """Open an xlsx workbook in read-only, values-only mode."""
    if not os.path.isfile(path):
        raise FileNotFoundError(f"Workbook not found: {path}")
    return load_workbook(path, read_only=True, data_only=data_only)


def workbook_sha256(path: str) -> str:
    """SHA-256 of the workbook file for provenance."""
    digest = hashlib.sha256()
    with open(path, "rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 256), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _row_is_data_header(cells: tuple[Any, ...]) -> bool:
    """True when a row carries the emission-factor header signature."""
    values = [clean_cell(c) for c in cells]
    return all(v in values for v in DATA_HEADER_SIGNATURE) and any(
        FACTOR_HEADER_RE.match(v) for v in values
    )


def detect_header_row(ws: Worksheet) -> Optional[tuple[int, dict[str, int]]]:
    """Locate the data header row and build a {column label: index} map.

    Returns ``(row_index_1based, column_map)`` or ``None`` when the worksheet
    carries no emission-factor table.
    """
    max_col = ws.max_column or 0
    for row_idx, row in enumerate(
        ws.iter_rows(
            min_row=1,
            max_row=min(MAX_HEADER_SCAN_ROWS, ws.max_row or 1),
            values_only=True,
        ),
        start=1,
    ):
        cells = tuple(row)
        if not _row_is_data_header(cells[:max_col]):
            continue
        column_map: dict[str, int] = {}
        for col_idx, value in enumerate(cells[:max_col], start=1):
            label = clean_cell(value)
            if label:
                column_map[label] = col_idx
        return row_idx, column_map
    return None


def parse_reporting_year(column_map: dict[str, int]) -> Optional[int]:
    """Extract the reporting year from the factor-column header."""
    for label in column_map:
        match = FACTOR_HEADER_RE.match(label)
        if match:
            return int(match.group(1))
    return None


def _sheet_is_documentation(ws: Worksheet) -> bool:
    """A worksheet is documentation when it carries label/value pairs but no data header."""
    if detect_header_row(ws):
        return False
    for row in ws.iter_rows(
        max_row=min(MAX_META_SCAN_ROWS, ws.max_row or 1),
        max_col=MAX_META_SCAN_COLS,
        values_only=True,
    ):
        values = [clean_cell(c) for c in row]
        for value in values:
            if any(value.startswith(suffix) for suffix in META_LABEL_SUFFIXES):
                return True
    return False


def scan_documentation_values(ws: Worksheet) -> dict[str, str]:
    """Collect label/value pairs from a documentation worksheet."""
    found: dict[str, str] = {}
    for row in ws.iter_rows(
        max_row=min(MAX_META_SCAN_ROWS, ws.max_row or 1),
        max_col=MAX_META_SCAN_COLS,
        values_only=True,
    ):
        values = [clean_cell(c) for c in row]
        for idx, value in enumerate(values):
            for suffix in META_LABEL_SUFFIXES:
                if not value.startswith(suffix):
                    continue
                label = suffix.rstrip(":").lower().replace(" ", "_")
                if suffix == "Produced by":
                    label = "producer"
                if suffix == "For technical queries":
                    label = "contact"
                following = [v for v in values[idx + 1 :] if v]
                if following:
                    found[label] = following[0]
                break
    return found


# ---------------------------------------------------------------------------
# Workbook analysis / discovery
# ---------------------------------------------------------------------------
def analyze_workbook(path: str) -> tuple[Workbook, WorkbookAnalysis]:
    """Open a workbook and classify every worksheet.

    Returns ``(workbook, analysis)``. Every worksheet is read and classified as
    ``data``, ``documentation`` or ``unsupported``; nothing is dropped.
    """
    wb = open_workbook(path)
    sheet_infos: list[WorksheetInfo] = []
    data_sheet_names: list[str] = []
    reporting_year: Optional[int] = None
    meta_values: dict[str, str] = {}

    for ws in wb.worksheets:
        header = detect_header_row(ws)
        max_row = ws.max_row
        max_col = ws.max_column
        if header is not None:
            row_idx, column_map = header
            year = parse_reporting_year(column_map)
            sheet_infos.append(
                WorksheetInfo(
                    name=ws.title,
                    sheet_type=SHEET_DATA,
                    max_row=max_row,
                    max_col=max_col,
                    header_row=row_idx,
                    columns=tuple(sorted(column_map.items(), key=lambda item: item[1])),
                    data_row_count=max(0, (max_row or 0) - row_idx),
                )
            )
            data_sheet_names.append(ws.title)
            if year is not None and reporting_year is None:
                reporting_year = year
        elif _sheet_is_documentation(ws):
            sheet_infos.append(
                WorksheetInfo(
                    name=ws.title,
                    sheet_type=SHEET_DOCUMENTATION,
                    max_row=max_row,
                    max_col=max_col,
                    notes=("Documentation sheet (ignored for import, reported).",),
                )
            )
            meta_values.update(scan_documentation_values(ws))
        else:
            sheet_infos.append(
                WorksheetInfo(
                    name=ws.title,
                    sheet_type=SHEET_UNSUPPORTED,
                    max_row=max_row,
                    max_col=max_col,
                    notes=("Unsupported sheet (no emission-factor table detected).",),
                )
            )

    year_meta: Optional[int] = None
    raw_year = meta_values.get("year", "")
    if raw_year.isdigit():
        year_meta = int(raw_year)

    workbook_meta = WorkbookMeta(
        source_path=path,
        file_sha256=workbook_sha256(path) if os.path.isfile(path) else "",
        file_size_bytes=os.path.getsize(path) if os.path.isfile(path) else None,
        title=meta_values.get("title", ""),
        version=meta_values.get("version", ""),
        year=year_meta,
        status=meta_values.get("status", ""),
        next_publication_date=meta_values.get("next_publication_date", ""),
        factor_set_label=meta_values.get("factor_set", ""),
        producer=meta_values.get("producer", ""),
        contact=meta_values.get("contact", ""),
    )

    if reporting_year is None and year_meta is not None:
        reporting_year = year_meta

    analysis = WorkbookAnalysis(
        meta=workbook_meta,
        worksheets=tuple(sheet_infos),
        data_sheet_names=tuple(data_sheet_names),
        reporting_year=reporting_year,
    )
    return wb, analysis



# ---------------------------------------------------------------------------
# Row extraction
# ---------------------------------------------------------------------------
def iter_data_rows(ws: Worksheet, info: WorksheetInfo) -> Iterator[tuple[int, tuple[Any, ...]]]:
    """Yield ``(row_number_1based, cells)`` for every row below the header."""
    if info.header_row is None:
        return
    for row_idx, row in enumerate(
        ws.iter_rows(
            min_row=info.header_row + 1,
            max_col=info.max_col or 0,
            values_only=True,
        ),
        start=info.header_row + 1,
    ):
        yield row_idx, tuple(row)


def parse_worksheet(ws: Worksheet, info: WorksheetInfo) -> tuple[list[ParsedRow], dict[str, int]]:
    """Parse every usable data row of one data worksheet.

    Returns ``(parsed_rows, counters)``. Counters cover rows scanned, blank
    rows, end markers, rows with an ID and rows with a factor value so the
    validation report can account for every row (nothing silently skipped).
    """
    column_map = dict(info.columns)
    factor_label: Optional[str] = next(
        (label for label in column_map if FACTOR_HEADER_RE.match(label)), None
    )
    if factor_label is None:
        raise ValueError(
            f"Worksheet {info.name!r} has no 'GHG Conversion Factor <year>' column."
        )

    def cell(label: str, cells: tuple[Any, ...]) -> str:
        idx = column_map.get(label)
        if idx is None or idx - 1 >= len(cells):
            return ""
        return clean_cell(cells[idx - 1])

    def raw_cell(label: str, cells: tuple[Any, ...]) -> Any:
        idx = column_map.get(label)
        if idx is None or idx - 1 >= len(cells):
            return None
        return cells[idx - 1]

    rows: list[ParsedRow] = []
    counters: dict[str, int] = {
        "rows_scanned": 0,
        "empty_rows": 0,
        "end_marker_rows": 0,
        "rows_parsed": 0,
        "rows_with_id": 0,
        "factors_with_value": 0,
    }

    for row_idx, cells in iter_data_rows(ws, info):
        counters["rows_scanned"] += 1
        if not any(c is not None and str(c).strip() for c in cells):
            counters["empty_rows"] += 1
            continue
        # DEFRA terminates factor tables with a row whose scope cell reads END.
        scope_idx = column_map.get("Scope")
        if scope_idx is not None and clean_cell(cells[scope_idx - 1]).upper() == "END":
            counters["end_marker_rows"] += 1
            continue

        parsed = ParsedRow(
            sheet_name=info.name,
            row_number=row_idx,
            defra_id=cell("ID", cells),
            scope=cell("Scope", cells),
            level1=cell("Level 1", cells),
            level2=cell("Level 2", cells),
            level3=cell("Level 3", cells),
            level4=cell("Level 4", cells),
            column_text=cell("Column Text", cells),
            uom=cell("UOM", cells),
            ghg_unit=cell("GHG/Unit", cells),
            factor_raw=raw_cell(factor_label, cells),
        )
        rows.append(parsed)
        counters["rows_parsed"] += 1
        if parsed.defra_id:
            counters["rows_with_id"] += 1
        if parsed.factor_raw is not None and str(parsed.factor_raw).strip():
            counters["factors_with_value"] += 1

    return rows, counters


# ---------------------------------------------------------------------------
# pandas-based sheet statistics
# ---------------------------------------------------------------------------
def sheet_dataframe(parsed_rows: list[ParsedRow]) -> pd.DataFrame:
    """Build a pandas DataFrame of a data sheet for analysis/statistics."""
    return pd.DataFrame([row.as_dict() for row in parsed_rows])


def pandas_sheet_stats(parsed_rows: list[ParsedRow]) -> dict[str, Any]:
    """Per-sheet statistics computed with pandas (feeds the statistics report)."""
    if not parsed_rows:
        return {"rows": 0, "columns": {}}
    df = sheet_dataframe(parsed_rows)
    columns: dict[str, Any] = {}
    for column in df.columns:
        series = df[column]
        columns[column] = {
            "non_null": int(series.notna().sum()),
            "null": int(series.isna().sum()),
            "unique": int(series.nunique()),
        }
    return {
        "rows": int(len(df)),
        "columns": columns,
        "scope_distribution": {
            str(key): int(value)
            for key, value in df["scope"].fillna("(missing)").value_counts().items()
        },
    }

