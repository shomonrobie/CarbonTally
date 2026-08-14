"""SEAI workbook parser (CarbonTally SEAI 2025 provider).

Reads the SEAI workbook with ``openpyxl`` (``data_only=True`` so cached formula
values are used — every value on the main sheet is an INDEX/MATCH + arithmetic
formula). Only the **authoritative sheet** ``Conversion and emission factors``
generates emission-factor rows (single source of truth rule); the timeseries,
blend and GHG sheets are reference data and are only reported.

Worksheet classification:
* ``Conversion and emission factors`` -> the authoritative data sheet.
* ``QAQC`` -> documentation (version history, author).
* All other sheets -> reference (never a factor source).

Layout notes (verified against the workbook):
* Column B holds the section header (``Liquid``/``Solid``/``Gas``/``Electricity``)
  and each row's activity name.
* Row pairs 19/20, 42/43, 57/58, 62/63 define the per-section headers;
  section header text rows (e.g. ``Petroleum``, ``Biofuel / bioliquid``,
  ``Blended petroleum & biofuel``, ``Fossil fuel``, ``Biomass``) are skipped.
* Emission-factor columns: F = gCO2/kWh, G = gCO2/MJ, H = kgCO2/kg
  (liquids/solids) or kgCO2/m^3 (gas), I = kgCO2/l (liquids).
* Conversion columns: C = toe/t, D = MJ/kg, E = MJ/l, J = kg/m^3,
  K = l/t, L = PE factor, M = Note, N = Year.
"""
from __future__ import annotations

import hashlib
import logging
from pathlib import Path
from typing import Optional

import openpyxl

from .models import (
    SeaiParsedRow,
    SeaiWorkbookData,
    SeaiWorksheetInfo,
    SeaiWorkbookMeta,
    _dec,
)

logger = logging.getLogger("seai_importer")

AUTHORITATIVE_SHEET = "Conversion and emission factors"
QAQC_SHEET = "QAQC"
REFERENCE_SHEETS = {
    "Energy content timeseries",
    "Emission factors timeseries",
    "Density timeseries",
    "Primary energy timeseries",
    "road_petrol_blend",
    "road_diesel_blend",
    "GHG_elec",
}

#: Top-level section header cells (column B) on the authoritative sheet.
_SECTION_HEADERS = {"Liquid", "Solid", "Gas", "Electricity"}

#: Sub-section header cells (column B) that are not data rows.
_SUBSECTION_HEADERS = {
    "Petroleum",
    "Biofuel / bioliquid",
    "Blended petroleum & biofuel",
    "Fossil fuel",
    "Biomass",
}

#: Column letter indexes used by the parser (1-based openpyxl indexes).
COL = {
    "B": 2, "C": 3, "D": 4, "E": 5, "F": 6, "G": 7,
    "H": 8, "I": 9, "J": 10, "K": 11, "L": 12, "M": 13, "N": 14,
}

def workbook_sha256(path: Path) -> str:
    """Return the SHA-256 hex digest of the workbook file."""
    h = hashlib.sha256()
    with open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def open_workbook(path: Path):
    """Load the workbook with cached values (``data_only=True``).

    ``read_only=True`` is used for analysis speed; it only exposes row-iteration
    access, which is all the parser needs.
    """
    return openpyxl.load_workbook(path, data_only=True, read_only=True)


def _parse_qaqc(sheet) -> SeaiWorkbookMeta:
    """Extract version/author metadata from the QAQC sheet."""
    version = ""
    status = ""
    contact = ""
    for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row or 60, 60), max_col=4):
        for cell in row:
            if cell.value is None or cell.row < 3:
                continue
            text = str(cell.value).strip()
            if not text:
                continue
            if text.startswith("V") and len(text) <= 12 and len(text) > len(version):
                version = text
            low = text.lower()
            if "status" in low and len(text) < 40 and ":" in text:
                status = text.split(":", 1)[1].strip()
            if "e-mail" in low or "email" in low or "@" in text:
                contact = text
    return SeaiWorkbookMeta(
        source_path="",
        title="Energy conversion and emission factors",
        latest_version=version,
        latest_version_date="",
        status=status,
        producer="SEAI Energy Statistics Team",
        contact=contact,
    )


def _classify_sheets(wb) -> tuple[tuple[SeaiWorksheetInfo, ...], str]:
    infos = []
    for ws in wb.worksheets:
        if ws.title == AUTHORITATIVE_SHEET:
            sheet_type = "data"
        elif ws.title == QAQC_SHEET:
            sheet_type = "documentation"
        else:
            sheet_type = "reference"
        infos.append(
            SeaiWorksheetInfo(
                name=ws.title,
                sheet_type=sheet_type,
                max_row=ws.max_row,
                max_col=ws.max_column,
            )
        )
    return tuple(infos), AUTHORITATIVE_SHEET


def _cell(row, col: int):
    """Safely read a cell value from an openpyxl row tuple."""
    if 1 <= col <= len(row):
        return row[col - 1].value
    return None


def _parse_authoritative_sheet(sheet) -> list[SeaiParsedRow]:
    """Extract the 28 data rows from the authoritative sheet (rows 19-69)."""
    rows: list[SeaiParsedRow] = []
    current_section: Optional[str] = None
    for offset, row in enumerate(sheet.iter_rows(min_row=19, max_row=69, max_col=14), start=19):
        r = offset
        name = _cell(row, COL["B"])
        if name is None:
            continue
        name_text = str(name).strip()
        if name_text in _SECTION_HEADERS:
            current_section = name_text
            continue
        if name_text in _SUBSECTION_HEADERS:
            continue
        if current_section is None:
            continue
        year_val = _cell(row, COL["N"])
        year = int(year_val) if isinstance(year_val, (int, float)) else None
        rows.append(
            SeaiParsedRow(
                row_number=r,
                name=name_text,
                top_section=current_section,
                toe_per_t=_dec(_cell(row, COL["C"])),
                mj_per_kg=_dec(_cell(row, COL["D"])),
                mj_per_l=_dec(_cell(row, COL["E"])),
                gco2_per_kwh=_dec(_cell(row, COL["F"])),
                gco2_per_mj=_dec(_cell(row, COL["G"])),
                kgco2_per_mass_unit=_dec(_cell(row, COL["H"])),
                kgco2_per_l=_dec(_cell(row, COL["I"])),
                density_kg_m3=_dec(_cell(row, COL["J"])),
                specific_vol_l_per_t=_dec(_cell(row, COL["K"])),
                pe_factor=_dec(_cell(row, COL["L"])),
                note=str(_cell(row, COL["M"]) or "").strip(),
                year=year,
            )
        )
    return rows


def analyze_workbook(path: Path) -> SeaiWorkbookData:
    """Parse the workbook and return the full discovered dataset."""
    if not path.exists():
        raise FileNotFoundError(f"SEAI workbook not found: {path}")
    wb = open_workbook(path)
    infos, authoritative = _classify_sheets(wb)
    if QAQC_SHEET in wb.sheetnames:
        qa = _parse_qaqc(wb[QAQC_SHEET])
    else:
        qa = SeaiWorkbookMeta(source_path="")
    meta = SeaiWorkbookMeta(
        source_path=str(path),
        file_sha256=workbook_sha256(path),
        file_size_bytes=path.stat().st_size,
        title=qa.title,
        reporting_year=None,
        latest_version=qa.latest_version,
        latest_version_date=qa.latest_version_date,
        status=qa.status,
        producer=qa.producer,
        contact=qa.contact,
    )
    rows = tuple(_parse_authoritative_sheet(wb[AUTHORITATIVE_SHEET]))
    return SeaiWorkbookData(
        meta=meta,
        worksheets=infos,
        authoritative_sheet=authoritative,
        rows=rows,
    )


def parse_worksheet(path: Path) -> SeaiWorkbookData:
    """Alias of :func:`analyze_workbook` (parser entry point)."""
    return analyze_workbook(path)

